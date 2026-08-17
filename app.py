import streamlit as st

# Trik agar Google bisa membaca file verifikasi HTML di Streamlit
query_params = st.query_params
if "google4ceece64915feaf4.html" in query_params: # ganti sesuai nama file milikmu
    st.write("google-site-verification: google4ceece64915feaf4.html")
    st.stop()
import os
import re
import json
import hashlib
import html
import logging
from pathlib import Path
from datetime import datetime
from urllib.parse import quote

import streamlit as st
import pandas as pd
import feedparser
import requests
from bs4 import BeautifulSoup
import plotly.express as px
from google import genai
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from datetime import datetime

# ============================================================
# KONFIGURASI HALAMAN & GOOGLE VERIFICATION
# ============================================================

st.set_page_config(
    page_title="Monitoring Berita Ekonomi Lamongan",
    page_icon="📰",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Tag Verifikasi Google + Styling CSS Dashboard Rapi
st.markdown("""
<meta name="google-site-verification" content="xrwK_BByxvJAfptvhoOoeWNHSvdb4vcGkTLxIz8k3ls" />
<style>
main {
    background-color: #f7f9fc;
}
.block-container {
    padding-top: 1.5rem;
    padding-bottom: 2rem;
}
.dashboard-title {
    font-size: 32px;
    font-weight: 700;
    color: #1f2937;
    margin-bottom: 0px;
}
.dashboard-subtitle {
    font-size: 15px;
    color: #6b7280;
    margin-bottom: 20px;
}
.kpi-card {
    background: white;
    padding: 20px;
    border-radius: 14px;
    border: 1px solid #e5e7eb;
    box-shadow: 0px 2px 8px rgba(0,0,0,0.04);
}
.kpi-title {
    font-size: 14px;
    color: #6b7280;
}
.kpi-value {
    font-size: 28px;
    font-weight: 700;
    color: #111827;
}
.section-title {
    font-size: 21px;
    font-weight: 650;
    margin-top: 20px;
    margin-bottom: 10px;
}
.news-card {
    background: white;
    padding: 16px;
    border-radius: 12px;
    border: 1px solid #e5e7eb;
    margin-bottom: 10px;
}
.news-title {
    font-size: 17px;
    font-weight: 650;
}
.news-meta {
    font-size: 13px;
    color: #6b7280;
}
.news-summary {
    font-size: 14px;
    color: #374151;
    line-height: 1.5;
}
.sidebar-title {
    font-size: 20px;
    font-weight: 700;
}
</style>
""", unsafe_allow_html=True)

# ============================================================
# INISIALISASI GEMINI AI & FILE PATH
# ============================================================

BASE_DIR = Path(__file__).resolve().parent
DATA_FILE = BASE_DIR / "berita_lamongan.csv"
LOG_FILE = BASE_DIR / "app.log"

logging.basicConfig(
    filename=LOG_FILE,
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s: %(message)s"
)
logger = logging.getLogger(__name__)

GEMINI_API_KEY = st.secrets.get("GEMINI_API_KEY", os.getenv("GEMINI_API_KEY", ""))

client = None
if GEMINI_API_KEY:
    try:
        client = genai.Client(api_key=GEMINI_API_KEY)
    except Exception as e:
        logger.error(f"Gagal inisialisasi Gemini Client: {e}")

# ============================================================
# 17 SEKTOR LAPANGAN USAHA BPS
# ============================================================

SEKTOR = {
    "A": "Pertanian, Kehutanan, dan Perikanan",
    "B": "Pertambangan dan Penggalian",
    "C": "Industri Pengolahan",
    "D": "Pengadaan Listrik dan Gas",
    "E": "Pengadaan Air, Pengelolaan Sampah, Limbah dan Daur Ulang",
    "F": "Konstruksi",
    "G": "Perdagangan Besar dan Eceran; Reparasi Mobil dan Sepeda Motor",
    "H": "Transportasi dan Pergudangan",
    "I": "Penyediaan Akomodasi dan Makan Minum",
    "J": "Informasi dan Komunikasi",
    "K": "Jasa Keuangan dan Asuransi",
    "L": "Real Estat",
    "M,N": "Jasa Perusahaan",
    "O": "Administrasi Pemerintahan, Pertahanan dan Jaminan Sosial Wajib",
    "P": "Jasa Pendidikan",
    "Q": "Jasa Kesehatan dan Kegiatan Sosial",
    "R,S,T,U": "Jasa Lainnya"
}

SEKTOR_KEYWORDS = {
    "A": ["pertanian","petani","sawah","padi","jagung","cabai","cabe","bawang","tebu","perkebunan","perikanan","nelayan","ikan","tambak","udang","rumput laut","peternakan","sapi","kambing","ayam","panen","pupuk","benih","hasil tani","pertanian lamongan"],
    "B": ["pertambangan","tambang","galian","pasir","batu","mineral","galian c"],
    "C": ["industri","pabrik","manufaktur","produksi","pengolahan","pabrikasi","industri pengolahan","sentra industri"],
    "D": ["listrik","pln","gas","energi","pembangkit","kelistrikan"],
    "E": ["sampah","limbah","air bersih","daur ulang","pengelolaan sampah","air minum","persampahan"],
    "F": ["konstruksi","pembangunan","gedung","jalan","jembatan","infrastruktur","perumahan","renovasi","proyek","pembangunan jalan","pembangunan gedung"],
    "G": ["perdagangan","pasar","pedagang","toko","ritel","eceran","grosir","jual beli","harga","komoditas","dealer","otomotif","kendaraan","pasar tradisional","pasar modern","distributor"],
    "H": ["transportasi","angkutan","bus","truk","pelabuhan","logistik","ekspedisi","pergudangan","terminal","jasa pengiriman","kendaraan umum"],
    "I": ["hotel","penginapan","restoran","rumah makan","warung","kuliner","cafe","kafe","pariwisata","wisata","akomodasi","destinasi wisata"],
    "J": ["digital","internet","telekomunikasi","teknologi","aplikasi","startup","online","e-commerce","digitalisasi","marketplace","internet"],
    "K": ["bank","perbankan","kredit","pembiayaan","asuransi","keuangan","pajak","investasi","bpr","pinjaman","finansial","perbankan lamongan"],
    "L": ["properti","real estat","perumahan","rumah","tanah","apartemen","developer","pengembang properti"],
    "M,N": ["jasa perusahaan","konsultan","jasa bisnis","tenaga kerja","outsourcing","konsultan bisnis"],
    "O": ["pemerintah","pemkab","pemda","bupati","anggaran","apbd","kebijakan pemerintah","program pemerintah","dinas","pemerintahan","kebijakan ekonomi"],
    "P": ["sekolah","pendidikan","kampus","universitas","guru","siswa","mahasiswa","beasiswa","pelatihan"],
    "Q": ["kesehatan","rumah sakit","rsud","puskesmas","dokter","pasien","bpjs","obat","kesehatan masyarakat"],
    "R,S,T,U": ["jasa lainnya","hiburan","organisasi","sosial","budaya","kesenian","olahraga","salon"]
}

ISU_KEYWORDS = {
    "Harga dan Inflasi": ["harga","inflasi","deflasi","naik","turun","mahal","murah","komoditas","kenaikan harga"],
    "Perdagangan": ["perdagangan","pasar","pedagang","jual","beli","ritel","grosir","distributor"],
    "Pertanian": ["pertanian","petani","panen","padi","jagung","cabai","tebu","pupuk","hasil pertanian"],
    "Perikanan": ["nelayan","ikan","tambak","perikanan","udang","laut","hasil tangkapan"],
    "Industri": ["industri","pabrik","produksi","manufaktur","pengolahan"],
    "UMKM": ["umkm","usaha mikro","usaha kecil","usaha menengah","pelaku usaha"],
    "Investasi": ["investasi","investor","modal","penanaman modal"],
    "Ketenagakerjaan": ["tenaga kerja","pekerja","buruh","lowongan","pengangguran","pekerjaan"],
    "Infrastruktur": ["jalan","jembatan","infrastruktur","pembangunan","konstruksi"],
    "Keuangan": ["bank","kredit","pembiayaan","asuransi","keuangan","pajak"],
    "Pariwisata": ["wisata","pariwisata","hotel","restoran","kuliner"],
    "Ekonomi Digital": ["digital","online","e-commerce","aplikasi","teknologi","marketplace"],
    "Ekonomi Daerah": ["ekonomi lamongan","pertumbuhan ekonomi","pdrb","ekonomi daerah","perekonomian"]
}

ECONOMIC_KEYWORDS = sorted({
    keyword
    for keyword_list in list(SEKTOR_KEYWORDS.values()) + list(ISU_KEYWORDS.values())
    for keyword in keyword_list
})

MEDIA_SEARCH = {
    "KlikJatim.com": "Lamongan ekonomi site:klikjatim.com",
    "KOMPAS.com": "Lamongan ekonomi site:kompas.com",
    "Radar Lamongan": "Lamongan ekonomi site:radarlamongan.jawapos.com",
    "ANTARAJATIM": "Lamongan ekonomi site:jatim.antaranews.com",
    "detikJatim": "Lamongan ekonomi site:detik.com",
    "BeritaJatim": "Lamongan ekonomi site:beritajatim.com",
    "Surya": "Lamongan ekonomi site:surya.co.id",
    "Jawa Pos": "Lamongan ekonomi site:jawapos.com",
    "Tribun": "Lamongan ekonomi site:tribunnews.com",
    "Times Indonesia": "Lamongan ekonomi site:timesindonesia.co.id",
    "Kumparan": "Lamongan ekonomi site:kumparan.com",
    "Media lainnya": "Lamongan ekonomi"
}

# ============================================================
# HELPER FUNCTIONS & AI ANALYZER
# ============================================================

def google_news_rss(query):
    return f"https://news.google.com/rss/search?q={quote(query)}&hl=id&gl=ID&ceid=ID:id"

def clean_text(text):
    if not text:
        return ""
    text = BeautifulSoup(str(text), "html.parser").get_text(" ")
    return re.sub(r"\s+", " ", text).strip()

def contains_keyword(text, keyword):
    return re.search(r"(?<!\w)" + re.escape(keyword.lower()) + r"(?!\w)", str(text).lower()) is not None

def detect_media(entry):
    try:
        source = entry.get("source")
        if source and source.get("title"):
            return source.get("title")
    except Exception:
        pass

    link = entry.get("link", "").lower()
    mapping = {
        "klikjatim": "KlikJatim.com",
        "kompas": "KOMPAS.com",
        "radarlamongan": "Radar Lamongan",
        "antaranews": "ANTARAJATIM",
        "detik": "detikJatim",
        "beritajatim": "BeritaJatim",
        "surya": "Surya",
        "jawapos": "Jawa Pos",
        "tribunnews": "Tribun",
        "timesindonesia": "Times Indonesia"
    }
    for key, val in mapping.items():
        if key in link:
            return val
    return "Media lainnya"

def classify_issue_fallback(title, summary):
    text = (str(title) + " " + str(summary)).lower()
    scores = {issue: sum(1 for kw in kws if contains_keyword(text, kw)) for issue, kws in ISU_KEYWORDS.items()}
    best = max(scores, key=scores.get)
    return "Ekonomi Umum" if scores[best] == 0 else best

def classify_sector_fallback(title, summary):
    text = (str(title) + " " + str(summary)).lower()
    scores = {kode: sum(1 for kw in kws if contains_keyword(text, kw)) for kode, kws in SEKTOR_KEYWORDS.items()}
    best = max(scores, key=scores.get)
    return "A - Pertanian, Kehutanan, dan Perikanan" if scores[best] == 0 else f"{best} - {SEKTOR[best]}"

def analyze_news_with_ai(title, raw_summary):
    """
    Menggunakan Gemini AI untuk membaca berita, membuat ringkasan cerdas,
    serta mengklasifikasikan Sektor BPS & Isu Ekonomi.
    """
    if not client:
        return {
            "sektor": classify_sector_fallback(title, raw_summary),
            "isu": classify_issue_fallback(title, raw_summary),
            "ringkasan": raw_summary if raw_summary and raw_summary != title else f"Informasi berita mengenai {title} di Kabupaten Lamongan."
        }

    prompt = f"""
    Kamu adalah pakar analis ekonomi BPS.
    Analisis berita berikut:
    Judul Berita: {title}
    Teks Awal/Deskripsi: {raw_summary}

    Tugas:
    1. Buat RINGKASAN CERDAS dalam 2-3 kalimat yang menjelaskan inti isi berita (jangan cuma mengulang judul!).
    2. Tentukan Sektor Lapangan Usaha BPS dari daftar ini (Gunakan format 'KODE - Nama Sektor', misal 'A - Pertanian, Kehutanan, dan Perikanan'):
       - A - Pertanian, Kehutanan, dan Perikanan
       - B - Pertambangan dan Penggalian
       - C - Industri Pengolahan
       - D - Pengadaan Listrik dan Gas
       - E - Pengadaan Air, Pengelolaan Sampah, Limbah dan Daur Ulang
       - F - Konstruksi
       - G - Perdagangan Besar dan Eceran; Reparasi Mobil dan Sepeda Motor
       - H - Transportasi dan Pergudangan
       - I - Penyediaan Akomodasi dan Makan Minum
       - J - Informasi dan Komunikasi
       - K - Jasa Keuangan dan Asuransi
       - L - Real Estat
       - M,N - Jasa Perusahaan
       - O - Administrasi Pemerintahan, Pertahanan dan Jaminan Sosial Wajib
       - P - Jasa Pendidikan
       - Q - Jasa Kesehatan dan Kegiatan Sosial
       - R,S,T,U - Jasa Lainnya
    3. Tentukan Isu Ekonomi Utama (Misal: UMKM, Harga dan Inflasi, Pertanian, Perdagangan, Investasi, Infrastruktur, Ketenagakerjaan, Ekonomi Daerah, dll).

    Balas HANYA dalam format JSON valid persis seperti ini:
    {{
        "sektor": "KODE - Nama Sektor",
        "isu": "Nama Isu Ekonomi",
        "ringkasan": "Isi ringkasan 2-3 kalimat"
    }}
    """
    try:
        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=prompt,
        )
        text_resp = response.text.strip()
        if "```json" in text_resp:
            text_resp = text_resp.split("```json")[1].split("```")[0].strip()
        elif "```" in text_resp:
            text_resp = text_resp.split("```")[1].split("```")[0].strip()
        
        data = json.loads(text_resp)
        return {
            "sektor": data.get("sektor", classify_sector_fallback(title, raw_summary)),
            "isu": data.get("isu", classify_issue_fallback(title, raw_summary)),
            "ringkasan": data.get("ringkasan", raw_summary)
        }
    except Exception as e:
        logger.error(f"Gemini API Error: {e}")
        return {
            "sektor": classify_sector_fallback(title, raw_summary),
            "isu": classify_issue_fallback(title, raw_summary),
            "ringkasan": raw_summary if raw_summary and raw_summary != title else f"Informasi berita mengenai {title} di Kabupaten Lamongan."
        }

def is_relevant(title, summary):
    text = (str(title) + " " + str(summary)).lower()
    lamongan_words = ["lamongan", "babat", "brondong", "paciran", "solokuro", "pucuk", "mantup", "sugio", "ngimbang", "kembangbahu", "sambeng", "kedungpring", "bluluk", "sukodadi", "tikung", "karangbinangun", "glagah", "deket", "turi", "maduran", "sekaran", "laren", "karanggeneng", "kalitengah"]
    has_location = any(contains_keyword(text, word) for word in lamongan_words)
    has_economic = any(contains_keyword(text, word) for word in ECONOMIC_KEYWORDS)
    return has_location and has_economic

def get_date(entry):
    try:
        if entry.get("published_parsed"):
            return datetime(*entry.published_parsed[:6]).strftime("%Y-%m-%d")
    except Exception:
        pass
    return datetime.now().strftime("%Y-%m-%d")

def get_summary(entry):
    summary = entry.get("summary") or entry.get("description") or ""
    summary = clean_text(summary)
    return (summary[:447] + "...") if len(summary) > 450 else summary

def make_id(title, link):
    return hashlib.md5((str(title) + str(link)).encode("utf-8")).hexdigest()

# ============================================================
# AMBIL BERITA
# ============================================================

def fetch_news():
    all_news = []
    fetch_errors = []
    progress = st.progress(0)
    status = st.empty()
    items = list(MEDIA_SEARCH.items())
    total = len(items)

    for i, (media_name, query) in enumerate(items):
        status.info(f"🔎 Mengambil & menganalisis berita dari {media_name}...")
        try:
            rss_url = google_news_rss(query)
            response = requests.get(rss_url, timeout=20, headers={"User-Agent": "Mozilla/5.0"})
            response.raise_for_status()
            feed = feedparser.parse(response.content)

            for entry in feed.entries:
                title = clean_text(entry.get("title", ""))
                link = entry.get("link", "")
                raw_summary = get_summary(entry)

                if not title or not link or not is_relevant(title, raw_summary):
                    continue

                detected_media = detect_media(entry)
                if detected_media == "Media lainnya":
                    detected_media = media_name

                # Proses analisis AI Gemini
                ai_res = analyze_news_with_ai(title, raw_summary)

                record = {
                    "ID": make_id(title, link),
                    "Tanggal Berita": get_date(entry),
                    "Media": detected_media,
                    "Judul Berita": title,
                    "Isu Ekonomi": ai_res["isu"],
                    "Sektor": ai_res["sektor"],
                    "Ringkasan Berita": ai_res["ringkasan"],
                    "Link Berita": link
                }
                all_news.append(record)
        except Exception:
            logger.exception(f"Gagal mengambil berita dari {media_name}")
            fetch_errors.append(media_name)

        progress.progress((i + 1) / total)

    status.success("✅ Pengambilan & analisis berita selesai.")
    progress.empty()

    if fetch_errors:
        st.warning("Sebagian sumber berita gagal diakses: " + ", ".join(fetch_errors))

    if not all_news:
        return pd.DataFrame()

    df = pd.DataFrame(all_news).drop_duplicates(subset=["ID"]).sort_values("Tanggal Berita", ascending=False)
    return df

def create_sample_data():
    sample = [
        {"ID":"1","Tanggal Berita":"2026-08-04","Media":"ANTARAJATIM","Judul Berita":"Aktivitas Ekonomi Kabupaten Lamongan Terus Tumbuh","Isu Ekonomi":"Ekonomi Daerah","Sektor":"A - Pertanian, Kehutanan, dan Perikanan","Ringkasan Berita":"Pemerintah Kabupaten Lamongan mencatatkan tren pertumbuhan positif di sektor komoditas unggulan daerah. Produktivitas pertanian serta aktivitas pasar rakyat menjadi penggerak utama kestabilan harga kebutuhan pokok.","Link Berita":"https://jatim.antaranews.com/"},
        {"ID":"2","Tanggal Berita":"2026-08-03","Media":"KlikJatim.com","Judul Berita":"Harga Komoditas Pertanian di Lamongan Mengalami Perubahan","Isu Ekonomi":"Harga dan Inflasi","Sektor":"A - Pertanian, Kehutanan, dan Perikanan","Ringkasan Berita":"Pergerakan harga cabai dan bawang di tingkat petani Lamongan mengalami fluktuasi akibat curah hujan. Pemkab melakukan pasokan berkala untuk menjaga stabilitas pasokan bahan pangan di pasaran.","Link Berita":"https://klikjatim.com/"},
        {"ID":"3","Tanggal Berita":"2026-08-02","Media":"detikJatim","Judul Berita":"Perdagangan dan UMKM Lamongan Terus Berkembang","Isu Ekonomi":"UMKM","Sektor":"G - Perdagangan Besar dan Eceran; Reparasi Mobil dan Sepeda Motor","Ringkasan Berita":"Ratusan produk UMKM lokal Lamongan kini siap menembus pasar ritel modern melalui program pendampingan sertifikasi halal dan kemasan digital yang digagas oleh dinas setempat.","Link Berita":"https://www.detik.com/jatim/"},
        {"ID":"4","Tanggal Berita":"2026-08-01","Media":"KOMPAS.com","Judul Berita":"Pembangunan Infrastruktur Dorong Aktivitas Ekonomi Lamongan","Isu Ekonomi":"Infrastruktur","Sektor":"F - Konstruksi","Ringkasan Berita":"Perbaikan akses jalur ruas jalan utama kawasan pesisir Utara Lamongan dipercepat guna memperlancar arus distribusi barang logistik dan konektivitas pelabuhan perikanan.","Link Berita":"https://www.kompas.com/"},
        {"ID":"5","Tanggal Berita":"2026-07-31","Media":"Radar Lamongan","Judul Berita":"Potensi Industri Pengolahan Lamongan Terus Dikembangkan","Isu Ekonomi":"Industri","Sektor":"C - Industri Pengolahan","Ringkasan Berita":"Sentra pengolahan hasil perikanan Lamongan memperluas jangkauan rantai pasok ekspor olahan hasil laut. Diversifikasi produk turunan ikan menjadi fokus peningkatan efisiensi sektor industri kawasan.","Link Berita":"https://radarlamongan.jawapos.com/"},
        {"ID":"6","Tanggal Berita":"2026-07-30","Media":"BeritaJatim","Judul Berita":"Investasi Menjadi Perhatian dalam Pengembangan Ekonomi Lamongan","Isu Ekonomi":"Investasi","Sektor":"K - Jasa Keuangan dan Asuransi","Ringkasan Berita":"Kemudahan izin usaha kawasan industri terpadu berhasil memikat minat investor domestik. Penyerapan tenaga kerja lokal diproyeksikan meningkat seiring pembukaan lahan industri baru.","Link Berita":"https://beritajatim.com/"},
        {"ID":"7","Tanggal Berita":"2026-07-29","Media":"Surya","Judul Berita":"Sektor Pariwisata Lamongan Terus Dikembangkan","Isu Ekonomi":"Pariwisata","Sektor":"I - Penyediaan Akomodasi dan Makan Minum","Ringkasan Berita":"Pengembangan fasilitas wisata bahari dan desa wisata berbasis kuliner lokal terus dipercantik. Peningkatan volume kunjungan wisatawan memberikan dampak positif pada tingkat okupansi penginapan.","Link Berita":"https://surya.co.id/"}
    ]
    return pd.DataFrame(sample)

# ============================================================
# LOAD DATA & SESSION STATE
# ============================================================

if "data" not in st.session_state:
    if DATA_FILE.exists():
        try:
            st.session_state.data = pd.read_csv(DATA_FILE)
        except Exception:
            logger.exception("Gagal membaca DATA_FILE")
            st.session_state.data = create_sample_data()
    else:
        st.session_state.data = create_sample_data()

# ============================================================
# SIDEBAR
# ============================================================

with st.sidebar:
    st.markdown('<div class="sidebar-title">📰 Monitoring Berita</div>', unsafe_allow_html=True)
    st.caption("Ekonomi Kabupaten Lamongan")
    st.divider()

    st.markdown("### ⚙️ Pengaturan")
    if st.button("🔄 Ambil Berita Terbaru", use_container_width=True):
        new_data = fetch_news()
        if not new_data.empty:
            final = pd.concat([st.session_state.data, new_data], ignore_index=True)
            final = final.drop_duplicates(subset=["ID"]).sort_values("Tanggal Berita", ascending=False)
            st.session_state.data = final
            try:
                final.to_csv(DATA_FILE, index=False, encoding="utf-8-sig")
                st.success("Data berita berhasil diperbarui.")
            except OSError:
                st.error("Data berhasil diambil, tetapi gagal disimpan ke CSV.")
        else:
            st.warning("Tidak ada data baru dari sumber berita.")
        st.rerun()

    if st.button("🗑️ Reset Data", use_container_width=True):
        st.session_state.data = create_sample_data()
        if DATA_FILE.exists():
            try:
                DATA_FILE.unlink()
            except OSError:
                pass
        st.rerun()

    st.divider()
    st.markdown("### 🔎 Filter")

df = st.session_state.data.copy()
if df.empty:
    df = create_sample_data()

df["Tanggal Berita"] = pd.to_datetime(df["Tanggal Berita"], errors="coerce")

with st.sidebar:
    min_date = df["Tanggal Berita"].min().date()
    max_date = df["Tanggal Berita"].max().date()

    date_range = st.date_input("📅 Periode Berita", value=(min_date, max_date))
    selected_media = st.multiselect("🌐 Media", sorted(df["Media"].dropna().unique()))
    selected_sector = st.multiselect("🏭 Sektor Lapangan Usaha", sorted(df["Sektor"].dropna().unique()))
    selected_issue = st.multiselect("📊 Isu Ekonomi", sorted(df["Isu Ekonomi"].dropna().unique()))
    keyword = st.text_input("🔎 Cari berita", placeholder="Judul, isu, kata kunci...")

# ============================================================
# FILTER DATA LOGIC
# ============================================================

filtered = df.copy()
if len(date_range) == 2:
    filtered = filtered[(filtered["Tanggal Berita"].dt.date >= date_range[0]) & (filtered["Tanggal Berita"].dt.date <= date_range[1])]
if selected_media:
    filtered = filtered[filtered["Media"].isin(selected_media)]
if selected_sector:
    filtered = filtered[filtered["Sektor"].isin(selected_sector)]
if selected_issue:
    filtered = filtered[filtered["Isu Ekonomi"].isin(selected_issue)]
if keyword:
    search_text = keyword.lower()
    filtered = filtered[filtered[["Judul Berita", "Isu Ekonomi", "Sektor", "Ringkasan Berita"]].fillna("").astype(str).apply(lambda row: row.str.lower().str.contains(search_text, regex=False).any(), axis=1)]

# ============================================================
# MAIN DASHBOARD TAMPILAN
# ============================================================

st.markdown('<div class="dashboard-title">📰 MONITORING BERITA EKONOMI LAMONGAN</div>', unsafe_allow_html=True)
st.markdown('<div class="dashboard-subtitle">Dashboard monitoring pemberitaan ekonomi Kabupaten Lamongan berdasarkan isu ekonomi dan 17 lapangan usaha.</div>', unsafe_allow_html=True)

st.info("💡 Gunakan tombol **Ambil Berita Terbaru** pada sidebar untuk mengambil berita terbaru dari berbagai media.")

# KPI METRICS
c1, c2, c3, c4 = st.columns(4)
c1.metric("📰 Total Berita", f"{len(filtered):,}")
c2.metric("📅 Berita Hari Ini", f"{len(filtered[filtered['Tanggal Berita'].dt.date == datetime.now().date()]):,}")
c3.metric("🌐 Media", f"{filtered['Media'].nunique():,}")
c4.metric("🏭 Sektor Terpantau", f"{filtered['Sektor'].nunique():,}")

st.divider()

# GRAFIK DASHBOARD
if not filtered.empty:
    st.markdown('<div class="section-title">📊 Ringkasan Monitoring</div>', unsafe_allow_html=True)
    col1, col2 = st.columns(2)

    with col1:
        sector_df = filtered["Sektor"].value_counts().reset_index()
        sector_df.columns = ["Sektor", "Jumlah"]
        fig_sector = px.bar(sector_df, x="Jumlah", y="Sektor", orientation="h", text="Jumlah", title="Berita Berdasarkan 17 Lapangan Usaha")
        fig_sector.update_traces(textposition="outside")
        fig_sector.update_layout(height=600, yaxis={"categoryorder": "total ascending"})
        st.plotly_chart(fig_sector, use_container_width=True)

    with col2:
        media_df = filtered["Media"].value_counts().reset_index()
        media_df.columns = ["Media", "Jumlah"]
        fig_media = px.pie(media_df, names="Media", values="Jumlah", hole=0.45, title="Distribusi Berita Berdasarkan Media")
        fig_media.update_layout(height=600)
        st.plotly_chart(fig_media, use_container_width=True)

    st.markdown('<div class="section-title">📈 Tren Pemberitaan</div>', unsafe_allow_html=True)
    trend_df = filtered.groupby("Tanggal Berita").size().reset_index(name="Jumlah Berita")
    fig_trend = px.line(trend_df, x="Tanggal Berita", y="Jumlah Berita", markers=True, title="Tren Jumlah Berita Ekonomi Lamongan")
    fig_trend.update_layout(height=400)
    st.plotly_chart(fig_trend, use_container_width=True)

    col3, col4 = st.columns(2)
    with col3:
        issue_df = filtered["Isu Ekonomi"].value_counts().reset_index()
        issue_df.columns = ["Isu", "Jumlah"]
        fig_issue = px.bar(issue_df, x="Isu", y="Jumlah", text="Jumlah", title="Distribusi Isu Ekonomi")
        fig_issue.update_layout(height=450, xaxis_tickangle=-40)
        st.plotly_chart(fig_issue, use_container_width=True)

    with col4:
        top_sector = filtered["Sektor"].value_counts().head(5).reset_index()
        top_sector.columns = ["Sektor", "Jumlah"]
        fig_top = px.bar(top_sector, x="Jumlah", y="Sektor", orientation="h", text="Jumlah", title="5 Sektor Paling Banyak Diberitakan")
        fig_top.update_layout(height=450, yaxis={"categoryorder": "total ascending"})
        st.plotly_chart(fig_top, use_container_width=True)

# TABEL BERITA
st.divider()
st.markdown('<div class="section-title">📋 Monitoring Berita</div>', unsafe_allow_html=True)
st.caption(f"Menampilkan {len(filtered)} berita sesuai filter.")

if not filtered.empty:
    table_df = filtered.copy()
    table_df["Tanggal Berita"] = table_df["Tanggal Berita"].dt.strftime("%d-%m-%Y")
    table_df = table_df[["Tanggal Berita", "Media", "Judul Berita", "Isu Ekonomi", "Sektor", "Ringkasan Berita", "Link Berita"]]
    
    for column in table_df.columns:
        table_df[column] = table_df[column].map(lambda value: html.escape(str(value)))

    table_df["Link Berita"] = table_df["Link Berita"].apply(lambda x: f'<a href="{x}" target="_blank" rel="noopener">🔗 Baca</a>')
    st.markdown(table_df.to_html(escape=False, index=False), unsafe_allow_html=True)
else:
    st.warning("Tidak ada berita yang sesuai dengan filter.")

# ============================================================
# EKSPOR DATA EXCEL PROFESIONAL
# ============================================================

st.divider()

st.markdown(
    '<div class="section-title">📥 Ekspor Data Excel</div>',
    unsafe_allow_html=True
)


def export_excel(df):

    output = BytesIO()

    wb = Workbook()

    ws = wb.active
    ws.title = "Monitoring Berita"

    # =====================================================
    # JUDUL
    # =====================================================

    ws.merge_cells("A1:G1")

    ws["A1"] = (
        "MONITORING BERITA EKONOMI "
        "KABUPATEN LAMONGAN"
    )

    ws["A1"].font = Font(
        name="Calibri",
        size=16,
        bold=True,
        color="FFFFFF"
    )

    ws["A1"].fill = PatternFill(
        fill_type="solid",
        fgColor="005B96"
    )

    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    ws.row_dimensions[1].height = 30

    # =====================================================
    # TANGGAL EKSPOR
    # =====================================================

    ws.merge_cells("A2:G2")

    ws["A2"] = (
        "Diekspor pada : "
        + datetime.now().strftime("%d-%m-%Y %H:%M")
    )

    ws["A2"].font = Font(
        name="Calibri",
        italic=True,
        size=10,
        color="666666"
    )

    ws["A2"].alignment = Alignment(
        horizontal="left",
        vertical="center"
    )

    # =====================================================
    # HEADER
    # =====================================================

    headers = [
        "Tanggal Berita",
        "Media",
        "Judul Berita",
        "Isu Ekonomi",
        "Sektor",
        "Ringkasan Berita",
        "Link Berita"
    ]

    for col_num, header in enumerate(headers, start=1):

        cell = ws.cell(
            row=4,
            column=col_num
        )

        cell.value = header

        cell.font = Font(
            name="Calibri",
            size=11,
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="0077B6"
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    ws.row_dimensions[4].height = 30

    # =====================================================
    # DATA
    # =====================================================

    start_row = 5

    for row in df.itertuples(index=False):

        ws.append(list(row))

    # =====================================================
    # BORDER
    # =====================================================

    thin = Side(
        style="thin",
        color="D9D9D9"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for row in ws.iter_rows(
        min_row=4,
        max_row=ws.max_row,
        min_col=1,
        max_col=7
    ):

        for cell in row:

            cell.border = border

            if cell.row >= start_row:

                cell.font = Font(
                    name="Calibri",
                    size=10
                )

                # Kolom teks panjang
                if cell.column in [3, 6]:

                    cell.alignment = Alignment(
                        horizontal="left",
                        vertical="top",
                        wrap_text=True
                    )

                else:

                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="top",
                        wrap_text=True
                    )

    # =====================================================
    # HYPERLINK
    # =====================================================

    for row in range(
        start_row,
        ws.max_row + 1
    ):

        cell = ws[f"G{row}"]

        if cell.value:

            cell.hyperlink = str(cell.value)

            cell.style = "Hyperlink"

    # =====================================================
    # LEBAR KOLOM
    # =====================================================

    column_widths = {

        "A": 16,
        "B": 20,
        "C": 55,
        "D": 25,
        "E": 45,
        "F": 70,
        "G": 55

    }

    for column, width in column_widths.items():

        ws.column_dimensions[column].width = width

    # =====================================================
    # AUTO FILTER
    # =====================================================

    if ws.max_row >= 4:

        ws.auto_filter.ref = (
            f"A4:G{ws.max_row}"
        )

    # =====================================================
    # FREEZE HEADER
    # =====================================================

    ws.freeze_panes = "A5"

    # =====================================================
    # PRINT SETTINGS
    # =====================================================

    ws.sheet_view.showGridLines = False

    ws.page_setup.orientation = "landscape"

    ws.page_setup.fitToWidth = 1

    ws.page_setup.fitToHeight = 0

    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # =====================================================
    # SIMPAN EXCEL
    # =====================================================

    wb.save(output)

    output.seek(0)

    return output


# ============================================================
# SIAPKAN DATA UNTUK EXCEL
# ============================================================

if not filtered.empty:

    excel_df = filtered.copy()

    # --------------------------------------------------------
    # FORMAT TANGGAL
    # --------------------------------------------------------

    excel_df["Tanggal Berita"] = pd.to_datetime(
        excel_df["Tanggal Berita"],
        errors="coerce"
    )

    excel_df["Tanggal Berita"] = (
        excel_df["Tanggal Berita"]
        .dt.strftime("%d-%m-%Y")
    )

    # --------------------------------------------------------
    # URUTAN KOLOM
    # --------------------------------------------------------

    excel_df = excel_df[
        [
            "Tanggal Berita",
            "Media",
            "Judul Berita",
            "Isu Ekonomi",
            "Sektor",
            "Ringkasan Berita",
            "Link Berita"
        ]
    ]

    # --------------------------------------------------------
    # BUAT FILE EXCEL
    # --------------------------------------------------------

    file_excel = export_excel(
        excel_df
    )

    # --------------------------------------------------------
    # DOWNLOAD
    # --------------------------------------------------------

    st.download_button(
        label="📥 Download Excel",
        data=file_excel,
        file_name=(
            "Monitoring_Berita_Lamongan_"
            + datetime.now().strftime("%Y%m%d")
            + ".xlsx"
        ),
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        use_container_width=True
    )

else:

    st.info(
        "Belum ada data yang dapat diekspor."
    )
st.divider()

with st.expander("📚 Lihat 17 Lapangan Usaha"):
    sector_display = pd.DataFrame({"Kode": list(SEKTOR.keys()), "Lapangan Usaha": list(SEKTOR.values())})
    st.dataframe(sector_display, use_container_width=True, hide_index=True)

st.divider()
st.caption("Dashboard Monitoring Berita Ekonomi Kabupaten Lamongan | Prototype Miniproject Magang BPS Kabupaten Lamongan")
